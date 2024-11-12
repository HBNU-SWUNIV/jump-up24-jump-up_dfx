import pprint
import time
import os
import numpy as np
import cv2
import copy
import csv
import openpyxl
import pytesseract
import json

debug_path = "./test"


class TSR:
    def __init__(self, image: str, kernels: list, colors: list = None, padding_size: int = 20,
                 match_score: float = 0.99, img_log: bool = False, save_path: str = None, table_cnt: int = None) -> None:
        self.image = cv2.imread(image)
        self.match_results = list()
        self.kernels = kernels
        self.colors = colors
        self.match_score = match_score
        self.padding_size = padding_size
        self.img_log = img_log
        self.padded_image = None
        self.padded_original_image = None
        self.result_image = None
        self.cells = []
        self.merge_type = ["HMerge", "VMerge", "UMerge"]
        self.merge_cells = []

        self.test_match_results = list()
        self.save_path = save_path
        self.table_cnt = table_cnt

    def image_processing(self) -> None:
        img = cv2.cvtColor(self.image, cv2.COLOR_BGR2GRAY)

        img_height, img_width = img.shape

        img_bin = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 5, 5)

        img_bin = 255 - img_bin

        # 커널 길이 설정
        kernel_len_ver = img_height // 40
        kernel_len_hor = img_width // 40

        # 커널 선언
        ver_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_len_ver))
        hor_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_len_hor, 1))
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))

        ver_image = cv2.erode(img_bin, ver_kernel, iterations=3)

        vertical_lines = cv2.dilate(ver_image, ver_kernel, iterations=3)

        hor_image = cv2.erode(img_bin, hor_kernel, iterations=3)

        horizontal_lines = cv2.dilate(hor_image, hor_kernel, iterations=3)

        combine_vh_image = cv2.addWeighted(vertical_lines, 0.5, horizontal_lines, 0.5, 0)

        combine_vh_image = cv2.erode(~combine_vh_image, kernel, iterations=1)

        _, binary_image = cv2.threshold(combine_vh_image, 128, 255, cv2.THRESH_BINARY)

        # 세선화 추가 작업 필요
        thinning_image = cv2.erode(~binary_image, None, iterations=1)

        # 패딩 이미지 및 결과 이미지(BGR scale) 저장
        self.padded_image = np.pad(~thinning_image, self.padding_size, constant_values=255)
        self.padded_original_image = np.pad(img, self.padding_size, constant_values=255)
        self.result_image = cv2.cvtColor(copy.deepcopy(self.padded_image), cv2.COLOR_GRAY2BGR)


    def matching(self):
        #
        match_image = (~self.padded_image // 255)
        for idx, kernel in enumerate(self.kernels):
            results = cv2.matchTemplate(match_image, kernel['kernel'], cv2.TM_CCOEFF_NORMED)
            box_locations = np.where(results >= self.match_score)
            for box in zip(*box_locations[::-1]):
                start_x, start_y = box
                end_x, end_y = start_x + 6, start_y + 6
                cv2.rectangle(self.result_image, (start_x, start_y), (end_x, end_y), (255, 0, 0), 1)
                # cv2.rectangle(self.result_image, (start_x + 3, start_y + 3), (end_x - 3, end_y - 3), (255, 0, 0), 1)
                # cv2.putText(self.result_image, f"{start_x + 3}, {start_y + 3}", (start_x, start_y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 0, 0))
                match_result = {
                    "point": (start_x + 3, start_y + 3),
                    "shape": kernel["shape"]
                }
                self.match_results.append(match_result)

        sorted_match_results = sorted(self.match_results, key=lambda x: (x["point"][0], x["point"][1]))
        # self.save_to_csv_from_list(sorted_match_results)
        self.match_results = sorted_match_results
        self.test_match_results = list(map(lambda x: (x["point"][0], x["point"][1]), sorted_match_results))

        return

    def find_cell_list(self):
        # Y 축으로 정렬 배열 생성
        start_point = self.match_results[0]["point"]
        rows = list(filter(lambda x: x["point"][0] == start_point[0], self.match_results))
        columns = list(filter(lambda x: x["point"][1] == start_point[1], self.match_results))

        # rows에서 y좌표만 저장
        row_point = list(map(lambda x: x["point"][1], rows))
        column_point = list(map(lambda x: x["point"][0], columns))
        max_col_len = len(column_point)

        points = []
        tables = []

        # 세로 병합 추가 필요
        for idx, row in enumerate(rows):
            points.append([result for result in self.match_results if result["point"][1] == row["point"][1]])
            idx_r = 0

            for col_p in column_point:
                # print(str(idx_r), "Data: ", str(points[idx][idx_r]["point"]), str(column_point[idx_r]))
                print(len(points[idx]), idx_r)
                if len(points[idx]) == idx_r:
                    print("여기는 옴?")
                    print(points[idx][0]["point"], points[idx][1]["point"])
                    points[idx].append({"point": [col_p, points[idx][0]["point"][1]], "shape": "UMerge"})
                elif not points[idx][idx_r]["point"][0] == col_p:
                    points[idx].insert(idx_r, {"point": [col_p, points[idx][idx_r]["point"][1]], "shape": "UMerge"})
                    print(str(idx_r), "insert: ", str(points[idx][idx_r]["point"]), str(column_point[idx_r]))
                    # points[idx].insert(idx_r, 0)
                idx_r += 1

            print("==================================")
        del idx_r

        rows_len = len(points)
        cols_len = len(points[0])

        """
        여기부터 테스트 코드
        """
        rectangles = []

        rect_list = []

        # Find Rectangle
        for i in range(rows_len - 1):
            rect = []
            for j in range(cols_len - 1):
                rectangle = {
                    "top_left": points[i][j],
                    "top_right": points[i][j + 1],
                    "bottom_left": points[i + 1][j],
                    "bottom_right": points[i + 1][j + 1]
                }
                rectangles.append(rectangle)
                rect.append(rectangle)
            rect_list.append(rect)

        # Rectangle logs
        for rect in rectangles:
            # print(f"사각형: {rect}")
            cv2.rectangle(self.result_image, rect["top_left"]["point"], rect["bottom_right"]["point"], (0, 0, 255), 1)

        # Merge Filtering
        # 아래 부분은 함수화 필요함 # 세로 병합 추가 필요
        for i in range(len(rect_list)):
            for j in range(len(rect_list[0])):
                if rect_list[i][j]["top_left"]["shape"] == "UMerge":
                    # i - 1 bottom_left
                    """
                    위 셀의 top_left : rect_list[i - 1][j]["top_left"]
                    같은 셀의 bottom_left : rect_list[i][j]["bottom_left"]
                    """
                    is_h_merge = False
                    try:
                        # 해당 포인트에 위쪽 포인트 검사
                        if rect_list[i - 1][j]["top_left"]["shape"] in ["TShape180", "LShape0", "LShape90", "HMerge"]:
                            is_h_merge = True
                        # 해당 포인트에 아래쪽 포인트 검사
                    except IndexError:
                        print("Index Error")

                    try:
                        if rect_list[i][j]["bottom_left"]["shape"] in ["TShape0", "LShape180", "LShape270", "HMerge"]:
                            is_h_merge = True
                    except IndexError:
                        print("Index Error")

                    if is_h_merge:
                        rect_list[i][j]["top_left"]["shape"] = "HMerge"
                    else:
                        rect_list[i][j]["top_left"]["shape"] = "VMerge"

                if rect_list[i][j]["top_right"]["shape"] == "UMerge":
                    # i - 1 bottom_right
                    """
                    위 셀의 top_right : rect_list[i - 1][j]["top_right"]
                    같은 셀의 bottom_right : rect_list[i][j]["bottom_right"]
                    """
                    is_h_merge = False
                    try:
                        # 해당 포인트에 위쪽 포인트 검사
                        if rect_list[i - 1][j]["top_right"]["shape"] in ["TShape180", "LShape0", "LShape90", "HMerge"]:
                            is_h_merge = True
                            rect_list[i][j]["top_right"]["shape"] = "HMerge"
                        # 해당 포인트에 아래쪽 포인트 검사
                    except IndexError:
                        print("Index Error")

                    try:
                        if rect_list[i][j]["bottom_right"]["shape"] in ["TShape0", "LShape180", "LShape270", "HMerge"]:
                            is_h_merge = True
                            rect_list[i][j]["top_right"]["shape"] = "HMerge"
                    except IndexError:
                        print("Index Error")

                    if is_h_merge:
                        rect_list[i][j]["top_right"]["shape"] = "HMerge"
                    else:
                        rect_list[i][j]["top_right"]["shape"] = "VMerge"

                if rect_list[i][j]["bottom_left"]["shape"] == "UMerge":
                    # i + 1 top_left
                    """
                    같은 셀의 top_left : rect_list[i][j]["top_left"]
                    아래 셀의 bottom_left : rect_list[i + 1][j]["bottom_left"]
                    """
                    is_h_merge = False
                    try:
                        # 해당 포인트에 위쪽 포인트 검사
                        if rect_list[i][j]["top_left"]["shape"] in ["TShape180", "LShape0", "LShape90", "HMerge"]:
                            is_h_merge = True
                            rect_list[i][j]["bottom_left"]["shape"] = "HMerge"
                        # 해당 포인트에 아래쪽 포인트 검사
                    except IndexError:
                        print("Index Error")

                    try:
                        if rect_list[i + 1][j]["bottom_left"]["shape"] in ["TShape0", "LShape180", "LShape270", "HMerge"]:
                            is_h_merge = True
                            rect_list[i][j]["bottom_left"]["shape"] = "HMerge"
                    except IndexError:
                        print("Index Error")

                    if is_h_merge:
                        rect_list[i][j]["bottom_left"]["shape"] = "HMerge"
                    else:
                        rect_list[i][j]["bottom_left"]["shape"] = "VMerge"

                if rect_list[i][j]["bottom_right"]["shape"] == "UMerge":
                    # i + 1 top_right
                    """
                    같은 셀의 top_right : rect_list[i][j]["top_right"]
                    아래 셀의 bottom_right : rect_list[i + 1][j]["bottom_right"]
                    """
                    is_h_merge = False
                    try:
                        # 해당 포인트에 위쪽 포인트 검사
                        if rect_list[i][j]["top_right"]["shape"] in ["TShape180", "LShape0", "LShape90", "HMerge"]:
                            is_h_merge = True
                            rect_list[i][j]["bottom_right"]["shape"] = "HMerge"
                        # 해당 포인트에 아래쪽 포인트 검사
                    except IndexError:
                        print("Index Error")

                    try:
                        if rect_list[i + 1][j]["bottom_right"]["shape"] in ["TShape0", "LShape180", "LShape270", "HMerge"]:
                            is_h_merge = True
                            rect_list[i][j]["bottom_right"]["shape"] = "HMerge"
                    except IndexError:
                        print("Index Error")

                    if is_h_merge:
                        rect_list[i][j]["bottom_right"]["shape"] = "HMerge"
                    else:
                        rect_list[i][j]["bottom_right"]["shape"] = "VMerge"

        h_merge = {
            "start_idx": [-1, -1], "end_idx": [-1, -1], "start_cnt": 0
        }

        v_merge = {
            "start_idx": [-1, -1], "end_idx": [-1, -1], "start_cnt": 0
        }
        # 세로 병합 추가 필요
        for i in range(len(rect_list)):
            for j in range(len(rect_list[0])):
                corners = [
                    rect_list[i][j]["top_left"]["shape"],
                    rect_list[i][j]["top_right"]["shape"],
                    rect_list[i][j]["bottom_left"]["shape"],
                    rect_list[i][j]["bottom_right"]["shape"]
                ]

                h_idx = sum(1 for corner in corners if corner == "HMerge")

                # 좌우 병합
                if h_idx > 0 and h_merge["start_idx"][0] == -1:
                    h_merge["start_idx"] = [i, j]
                    h_merge["start_cnt"] = h_idx

                elif h_merge["start_idx"][0] != -1 and h_merge["start_cnt"] >= h_idx:
                    h_merge["end_idx"] = [i, j]
                    self.merge_cells.append(h_merge)
                    h_merge = {"start_idx": [-1, -1], "end_idx": [-1, -1], "start_cnt": 0}

                elif h_merge["end_idx"][0] != -1 and h_merge["start_cnt"] <= h_idx:
                    h_merge["end_idx"] = [i, j]
                    self.merge_cells.append(h_merge)
                    h_merge = {"start_idx": [i, j], "end_idx": [-1, -1], "start_cnt": h_idx}

        for i in range(len(rect_list[0])):
            for j in range(len(rect_list)):
                corners = [
                    rect_list[j][i]["top_left"]["shape"],
                    rect_list[j][i]["top_right"]["shape"],
                    rect_list[j][i]["bottom_left"]["shape"],
                    rect_list[j][i]["bottom_right"]["shape"]
                ]
                v_idx = sum(1 for corner in corners if corner == "VMerge")
                # 상하 병합

                if v_idx > 0 and v_merge["start_idx"][1] == -1:
                    print("여기 들어옴?")
                    v_merge["start_idx"] = [j, i]
                    v_merge["start_cnt"] = v_idx
                elif v_merge["start_idx"][1] != -1 and v_merge["start_cnt"] >= v_idx:
                    v_merge["end_idx"] = [j, i]
                    self.merge_cells.append(v_merge)
                    v_merge = {"start_idx": [-1, -1], "end_idx": [-1, -1], "start_cnt": 0}
                elif v_merge["end_idx"][1] != -1 and v_merge["start_cnt"] <= v_idx:
                    v_merge["end_idx"] = [j, i]
                    self.merge_cells.append(v_merge)
                    v_merge = {"start_idx": [j, i], "end_idx": [-1, -1], "start_cnt": v_idx}

        print(f"merge_cells : {self.merge_cells}")

        # self.cells에 병합 데이터 넣기
        # 상하 병합 넣기 작업 해야함
        for merge in self.merge_cells:
            print(merge["start_idx"], merge["end_idx"])
            # 좌우 병합만 테스트
            if merge["start_idx"][0] == merge["end_idx"][0]:
                for i in range(merge["start_idx"][1], merge["end_idx"][1] + 1):
                    rect_list[merge["start_idx"][0]][i]["merge"] = merge
                    # print(rect_list[merge["start_idx"][0]][i])
            elif merge["start_idx"][1] == merge["end_idx"][1]:
                for i in range(merge["start_idx"][0], merge["end_idx"][0] + 1):
                    rect_list[i][merge["start_idx"][1]]["merge"] = merge

        self.cells = rect_list

        # pprint.pprint(self.cells)

        return 0

    def find_cell_list_test(self):
        # Y 축으로 정렬 배열 생성
        start_point = self.test_match_results[0]
        rows = list(filter(lambda x: x[0] == start_point[0], self.test_match_results))
        columns = list(filter(lambda x: x[1] == start_point[1], self.test_match_results))
        points = []

        for row in rows:
            points.append([result for result in self.test_match_results if result[1] == row[1]])

        # rows에서 y좌표만 저장
        row_point = list(map(lambda x: x[1], rows))
        column_point = list(map(lambda x: x[0], columns))

        return

    def find_cell_for_gpt(self):
        img_height, img_width = self.padded_image.shape

        for idx, match in enumerate(self.match_results):
            cell = [match["point"]]
            start_point = match["point"]
            direction = format(match["shape"]['direction'], '04b')

            for i in range(4):
                if direction[i] == "1":
                    if i == 1:  # Down
                        for j in range(start_point[1] + 1, img_height):
                            down_match = next((k for k in self.match_results if k["point"] == (start_point[0], j)),
                                              None)
                            if down_match and format(down_match["shape"]["direction"], '04b')[0] == "1":
                                cell.append(down_match["point"])

                                for l in range(down_match["point"][0] + 1, img_width):
                                    right_match = next(
                                        (n for n in self.match_results if n["point"] == (l, down_match["point"][1])),
                                        None)
                                    if right_match and format(right_match["shape"]["direction"], '04b')[2] == "1":
                                        cell.append(right_match["point"])

                                        for m in range(right_match["point"][1] - 1, 0, -1):
                                            up_match = next((o for o in self.match_results if
                                                             o["point"] == (right_match["point"][0], m)), None)
                                            if up_match and format(up_match["shape"]["direction"], '04b')[1] == "1":
                                                cell.append(up_match["point"])
                                                break

                                        break

                                break

            print(len(cell), end=" ")
            if len(cell) == 2:
                print(cell, end=" ")
            if len(cell) == 4:
                print(cell, end=" ")
                cv2.rectangle(self.result_image, cell[0], cell[2], (0, 0, 255), 1)
                self.cells.append(cell)
            print(idx + 1)
        print(len(self.match_results))
        cv2.imwrite("cellfind_for_gpt.png", self.result_image)
        print(self.cells)

    def find_cell(self):
        img_height, img_width = self.padded_image.shape
        # print(img_height, img_width)
        for idx, match in enumerate(self.match_results):
            cell = [match["point"]]
            start_point = match["point"]
            direction = match["shape"]['direction']
            # print(start_point, format(direction, '04b'))
            # print(match["shape"]["shape"])
            # 방향별 탐색 순서는 상, 하, 좌, 우
            # index : (0, 상), (1, 하), (2, 좌), (3, 우)
            for i in range(4):
                # print(format(direction, '04b')[i])
                if format(direction, '04b')[i] == "1":
                    if i == 0:      # 상
                        pass

                    elif i == 1:    # 하
                        is_find = False
                        is_find2 = False
                        is_find3 = False
                        for j in range(start_point[1] + 1, img_height):
                            for k in self.match_results:
                                if k["point"][0] == start_point[0] and k["point"][1] == j:
                                    if format(k["shape"]["direction"], '04b')[0] == "1":
                                        cell.append(k["point"])
                                        for l in range(k["point"][0] + 1, img_width):
                                            for n in self.match_results:
                                                if n["point"][1] == k["point"][1] and n["point"][0] == l:
                                                    if format(n["shape"]["direction"], '04b')[2] == "1":
                                                        cell.append(n["point"])
                                                        for m in range(n["point"][1] - 1, 0, -1):
                                                            for o in self.match_results:
                                                                if o["point"][0] == n["point"][0] and o["point"][1] == m:
                                                                    if format(o["shape"]["direction"], '04b')[1] == "1":
                                                                        cell.append(o["point"])
                                                                        is_find3 = True
                                                                        break

                                                            if is_find3:
                                                                break

                                                        is_find2 = True
                                                        break
                                            if is_find2:
                                                break

                                    is_find = True
                                    break
                            if is_find:
                                break

                    elif i == 2:    # 좌
                        pass

                    elif i == 3:    # 우
                        pass

                # Reception 다 찾을 떄 탈출
            print(len(cell), end=" ")
            if len(cell) == 2:
                print(cell, end=" ")
            if len(cell) == 4:
                print(cell, end=" ")
                cv2.rectangle(self.result_image, (cell[0]), (cell[2]), (0, 0, 255), 1)
                self.cells.append(cell)
            print(idx + 1)
        print(len(self.match_results))
        cv2.imwrite("cellfind2.png", self.result_image)
        print(self.cells)

    def new_cells_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        config = "--oem 1 --psm 3"
        log_json = {
            "result": []
        }
        for idx_c, col in enumerate(self.cells):
            for idx_r, row in enumerate(col):
                x1, y1, x2, y2 = None, None, None, None
                merge_info = None
                if "merge" in row:
                    if "isContinue" in self.cells[idx_c][idx_r]["merge"]:
                        continue
                    else:
                        start_idx = row["merge"]["start_idx"]
                        end_idx = row["merge"]["end_idx"]
                        if start_idx[0] == end_idx[0]:
                            for i in range(start_idx[1], end_idx[1] + 1):
                                self.cells[start_idx[0]][i]["merge"]["isContinue"] = True
                        elif start_idx[1] == end_idx[1]:
                            for i in range(start_idx[0], end_idx[0] + 1):
                                self.cells[i][start_idx[1]]["merge"]["isContinue"] = True

                    i, j = row["merge"]["end_idx"][0], row["merge"]["end_idx"][1]
                    merge_info = {
                        "start_row": start_idx[0] + 1,
                        "start_column": start_idx[1] + 1,
                        "end_row": end_idx[0] + 1,
                        "end_column": end_idx[1] + 1
                    }
                    x1 = row["top_left"]["point"][0]
                    y1 = row["top_left"]["point"][1]
                    x2 = self.cells[i][j]["bottom_right"]["point"][0]
                    y2 = self.cells[i][j]["bottom_right"]["point"][1]
                else:
                    x1 = row["top_left"]["point"][0]
                    y1 = row["top_left"]["point"][1]
                    x2 = row["bottom_right"]["point"][0]
                    y2 = row["bottom_right"]["point"][1]

                img = self.padded_original_image[y1:y2, x1:x2]
                text = pytesseract.image_to_string(img, config=config)
                # text = "debug"
                # print(text)
                ws.cell(row=idx_c + 1, column=idx_r + 1, value=text)
                if merge_info is not None:
                    ws.merge_cells(start_row=merge_info["start_row"], start_column=merge_info["start_column"],
                                   end_row=merge_info["end_row"], end_column=merge_info["end_column"])

        wb.save(f"{self.save_path}/{self.table_cnt}_result.xlsx")
        return

    def cells_to_excel(self):
        columns = []
        column = []
        column_idx = self.cells[0][0][0]
        is_merge = False
        col_len = []

        for cell in self.cells:
            if column_idx == cell[0][0]:
                column.append(cell)
            else:
                columns.append(column)
                column_idx = cell[0][0]
                column = [cell]

        columns.append(column)
        col_max_len = len(columns[0])

        for col in columns:
            if len(col) > col_max_len:
                col_max_len = len(col)
                is_merge = True

        wb = openpyxl.Workbook()
        ws = wb.active
        config = ("-l eng --oem 1")
        for idx_c, col in enumerate(columns):
            print(idx_c, col)
            for idx_r, row in enumerate(col):
                print(idx_r, row)
                print(row[0][1], row[1][1], row[0][0], row[2][0])
                img = self.padded_original_image[row[0][1]:row[1][1], row[0][0]:row[2][0]]

                text = pytesseract.image_to_string(img, config='config')
                # print(text)
                #
                # cv2.imshow("Original Image", self.padded_original_image)
                # cv2.imshow("Original Image", img)
                # cv2.waitKey(0)
                # cv2.destroyAllWindows()
                cv2.imwrite(f"output/cells/cell{idx_c}_{idx_r}.png", img)
                ws.cell(row=idx_r + 1, column=idx_c + 1, value=text)

        # 병합셀 찾기
        for idx_c, col in enumerate(columns):
            for idx_r, row in enumerate(col):
                h, w = row[0][1], row[0][2]


        wb.save("result.xlsx")
        return

    def save_to_csv(self):
        with open(f'{debug_path}/data.csv', 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerow(['no', 'point', 'shape'])
            for idx, match in enumerate(self.match_results):
                csv_writer.writerow([idx+1, match['point'], match['shape']['shape']])

    def save_to_csv_from_list(self, result):
        with open(f'{debug_path}/data.csv', 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerow(['no', 'point', 'shape'])
            for idx, match in enumerate(result):
                csv_writer.writerow([idx+1, match['point'], match['shape']['shape']])

    def run(self):
        start_time = time.time()
        self.image_processing()
        image_processing_time = time.time() - start_time

        start_time = time.time()
        self.matching()
        matching_time = time.time() - start_time

        # self.find_cell()
        # start_time = time.time()
        # self.find_cell_for_gpt()
        # find_cell_for_gpt_time = time.time() - start_time

        start_time = time.time()
        self.find_cell_list()
        find_cell_list_time = time.time() - start_time

        # self.find_cell_list_test()
        self.new_cells_to_excel()

        print(f"image_processing execution time: {image_processing_time} seconds")
        print(f"matching execution time: {matching_time} seconds")
        print(f"find_cell_list execution time: {find_cell_list_time} seconds")
        pass
